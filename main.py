import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# loop line start row 2 -> last line
products_count_result = {}
total_value = {}
minimal_inventory = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value  # col 4e name
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_id = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)  # insert value to this cell

    # calculation number of products per supplier
    if supplier_name not in products_count_result:
        products_count_result[supplier_name] = 0

    products_count_result[supplier_name] += 1

    # calculation total value of inventory per supplier
    if supplier_name not in total_value:
        total_value[supplier_name] = 0

    total_value[supplier_name] += inventory * price

    # searching minimal inventory
    if inventory < 10:
        minimal_inventory[int(product_id)] = int(inventory)

    # insert value for total inventory price to excel
    inventory_price.value = inventory * price

print(products_count_result)
print(total_value)
print(minimal_inventory)

inv_file.save("inventory_with_total.xlsx")  # save in new file
